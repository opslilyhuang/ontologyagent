"""文件数据源适配器 — CSV / Excel / JSON / PDF / Word / TXT。"""
import csv
import json
from pathlib import Path
from typing import Any, AsyncIterator

import aiofiles

from .base import DataAdapter, TableSchema, ColumnInfo, DataRecord


class FileAdapter(DataAdapter):

    def __init__(self, config: dict[str, Any]):
        super().__init__(config)
        self.file_path = Path(config["file_path"])
        self.file_type = self.file_path.suffix.lower()

    async def test_connection(self) -> bool:
        return self.file_path.exists()

    # ── Schema 提取 ──────────────────────
    async def extract_schema(self) -> list[TableSchema]:
        if self.file_type in (".csv",):
            return await self._csv_schema()
        elif self.file_type in (".xlsx", ".xls"):
            return await self._excel_schema()
        elif self.file_type == ".json":
            return await self._json_schema()
        elif self.file_type in (".pdf", ".docx", ".txt"):
            # 非结构化文本用单列 "content" 模拟
            return [TableSchema(
                table_name=self.file_path.stem,
                columns=[ColumnInfo(name="content", original_type="text")],
                row_count=1,
            )]
        raise ValueError(f"Unsupported file type: {self.file_type}")

    async def _csv_schema(self) -> list[TableSchema]:
        import asyncio

        def _read():
            with open(self.file_path, newline="", encoding="utf-8") as f:
                reader = csv.reader(f)
                header = next(reader)
                rows = sum(1 for _ in reader)
            return header, rows

        header, rows = await asyncio.get_event_loop().run_in_executor(None, _read)
        columns = [ColumnInfo(name=col, original_type="varchar") for col in header]
        return [TableSchema(table_name=self.file_path.stem, columns=columns, row_count=rows)]

    async def _excel_schema(self) -> list[TableSchema]:
        import openpyxl
        import asyncio

        def _read():
            wb = openpyxl.load_workbook(self.file_path, read_only=True)
            schemas = []
            for ws in wb.worksheets:
                rows = list(ws.iter_rows(max_row=2, values_only=False))
                if not rows:
                    continue
                header = [str(cell.value or "") for cell in rows[0]]
                columns = [ColumnInfo(name=h, original_type="varchar") for h in header]
                schemas.append(TableSchema(table_name=ws.title, columns=columns))
            wb.close()
            return schemas

        return await asyncio.get_event_loop().run_in_executor(None, _read)

    async def _json_schema(self) -> list[TableSchema]:
        import asyncio

        def _read():
            with open(self.file_path, encoding="utf-8") as f:
                data = json.load(f)
            if isinstance(data, list) and data:
                keys = list(data[0].keys()) if isinstance(data[0], dict) else ["value"]
                return [TableSchema(
                    table_name=self.file_path.stem,
                    columns=[ColumnInfo(name=k, original_type="varchar") for k in keys],
                    row_count=len(data),
                )]
            elif isinstance(data, dict):
                return [TableSchema(
                    table_name=self.file_path.stem,
                    columns=[ColumnInfo(name=k, original_type=type(v).__name__) for k, v in data.items()],
                    row_count=1,
                )]
            return []

        return await asyncio.get_event_loop().run_in_executor(None, _read)

    # ── 采样 ─────────────────────────────
    async def sample_data(self, table: str, limit: int = 100) -> list[dict[str, Any]]:
        if self.file_type == ".csv":
            return await self._csv_sample(limit)
        elif self.file_type in (".xlsx", ".xls"):
            return await self._excel_sample(table, limit)
        elif self.file_type == ".json":
            return await self._json_sample(limit)
        elif self.file_type in (".pdf", ".docx", ".txt"):
            text = await self._read_text_content()
            return [{"content": text}]
        return []

    async def _csv_sample(self, limit: int) -> list[dict[str, Any]]:
        import asyncio

        def _read():
            with open(self.file_path, newline="", encoding="utf-8") as f:
                reader = csv.DictReader(f)
                return [row for i, row in enumerate(reader) if i < limit]

        return await asyncio.get_event_loop().run_in_executor(None, _read)

    async def _excel_sample(self, sheet: str, limit: int) -> list[dict[str, Any]]:
        import openpyxl
        import asyncio

        def _read():
            wb = openpyxl.load_workbook(self.file_path, read_only=True)
            ws = wb[sheet]
            rows = list(ws.iter_rows(max_row=limit + 1, values_only=True))
            if not rows:
                return []
            header = [str(v or "") for v in rows[0]]
            result = [dict(zip(header, [str(v) if v is not None else "" for v in row])) for row in rows[1:]]
            wb.close()
            return result

        return await asyncio.get_event_loop().run_in_executor(None, _read)

    async def _json_sample(self, limit: int) -> list[dict[str, Any]]:
        import asyncio

        def _read():
            with open(self.file_path, encoding="utf-8") as f:
                data = json.load(f)
            if isinstance(data, list):
                return data[:limit]
            return [data]

        return await asyncio.get_event_loop().run_in_executor(None, _read)

    async def _read_text_content(self) -> str:
        if self.file_type == ".txt":
            async with aiofiles.open(self.file_path, encoding="utf-8") as f:
                return await f.read()
        elif self.file_type == ".pdf":
            import pdfplumber
            import asyncio

            def _read():
                with pdfplumber.open(self.file_path) as pdf:
                    return "\n".join(page.extract_text() or "" for page in pdf.pages)

            return await asyncio.get_event_loop().run_in_executor(None, _read)
        elif self.file_type == ".docx":
            import docx
            import asyncio

            def _read():
                doc = docx.Document(self.file_path)
                return "\n".join(p.text for p in doc.paragraphs)

            return await asyncio.get_event_loop().run_in_executor(None, _read)
        return ""

    # ── 流式读取 ──────────────────────────
    async def stream_records(self, table: str) -> AsyncIterator[DataRecord]:
        samples = await self.sample_data(table, limit=999999)
        for row in samples:
            yield DataRecord(table=table, data=row)
